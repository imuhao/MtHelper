import sys
from PyQt6 import QtWidgets, QtCore,QtGui
from PyQt6.QtWidgets import QLabel,QFileDialog,QTreeWidgetItem
from ui.ui import Ui_MainWindow
from ui.number_widget import NumericTableWidgetItem
from datetime import datetime, timedelta
from PyQt6.QtCore import QDate,QFileSystemWatcher,Qt,QDateTime,QTimer
from PyQt6.QtGui import QStandardItemModel, QStandardItem
import urllib3
from data.request_data import RequestData
from data.constant import Contants
import qdarkstyle
from config import config
from openpyxl import Workbook
from utils import cookie_util

class MyApp(QtWidgets.QMainWindow):

    sign_chart = QtCore.pyqtSignal(dict)

    def __init__(self):
        super().__init__()
        
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.apply_stylesheet()
        self.setup_widget()
        self.requestData = RequestData()
        self.mouse_pos = None


        self.refresh_shop_data()


    def apply_stylesheet(self):
        QtWidgets.QApplication.setStyle("Fusion")
        QtWidgets.QApplication.setPalette(QtWidgets.QApplication.style().standardPalette())


    def setup_network(self):
        # 发起请求时忽略证书验证警告
        urllib3.disable_warnings()

    def setup_widget(self):
        """设置表格控件"""
        header = self.ui.tableWidget.horizontalHeader()
        self.ui.tableWidget.setColumnCount(9)

        self.ui.tableWidget.setColumnWidth(Contants.index_name, 200)
        self.ui.tableWidget.setColumnWidth(Contants.index_deal_amount_value, 200)
        self.ui.tableWidget.setColumnWidth(Contants.index_deal_relative_cycle, 100)
        self.ui.tableWidget.setColumnWidth(Contants.index_deal_ranking_interval, 100)
        self.ui.tableWidget.setColumnWidth(Contants.index_transaction_turnover, 100)
        self.ui.tableWidget.setColumnWidth(Contants.index_transaction_relative_cycle, 100)
        self.ui.tableWidget.setColumnWidth(Contants.index_average_price, 100)
        self.ui.tableWidget.setColumnWidth(Contants.index_refund_amount_value, 100)
        self.ui.tableWidget.setColumnWidth(Contants.index_refund_relative_cycle, 100)

        self.ui.purchaseTable.setColumnWidth(Contants.index_purchase_shop_name, 200)
        self.ui.purchaseTable.setColumnWidth(Contants.index_purchase_order_number, 70)
        self.ui.purchaseTable.setColumnWidth(Contants.index_purchase_add_time, 155)
        self.ui.purchaseTable.setColumnWidth(Contants.index_purchase_order_time, 155)
        self.ui.purchaseTable.setColumnWidth(Contants.index_purchase_turnover, 85)
        self.ui.purchaseTable.setColumnWidth(Contants.index_estimated_revenue, 85)
        self.ui.purchaseTable.setColumnWidth(Contants.index_purchase_cost, 85)
        self.ui.purchaseTable.setColumnWidth(Contants.index_purchase_profit, 85)
        self.ui.purchaseTable.setColumnWidth(Contants.index_profit_margin, 85)

        #刷新店铺数据
        self.ui.btnRefreshData.clicked.connect(self.refresh_shop_data)
        #刷新采购数据
        self.ui.btnRefreshPurchaseData.clicked.connect(self.refresh_purchase_data)
        self.ui.tabWidget.currentChanged.connect(self.on_tab_changed)
        #窗口按钮
        self.ui.btnClose.clicked.connect(self.on_close)
        self.ui.btnMini.clicked.connect(self.on_mini)
        #店铺订单数据
        self.ui.btnShopOrder.clicked.connect(self.on_shop_order_click_listener)
        #导出订单数据
        self.ui.btnExportOrder.clicked.connect(self.on_order_export_click_listener)
        #下拉选择框监听
        self.ui.cbShop.currentTextChanged.connect(self.selectionShopChanged)

        #设置订单查询下拉框时间
        # 获取当前时间
        now = datetime.now()

        # 获取今天零点的时间
        start_of_today = datetime(now.year, now.month, now.day, 0, 0, 0)

        # 获取今天23:59点
        end_of_today = datetime(now.year, now.month, now.day, 23, 59, 0)

        # 将 datetime 对象转换为 QDateTime 对象
        qdt_start_of_today = QDateTime(start_of_today.year, start_of_today.month, start_of_today.day, 
                                    start_of_today.hour, start_of_today.minute, start_of_today.second)
        qdt_end_of_today = QDateTime(end_of_today.year, end_of_today.month, end_of_today.day, 
                                    end_of_today.hour, end_of_today.minute, end_of_today.second)
        self.ui.dataStartOrder.setDateTime(qdt_start_of_today)
        self.ui.dataEndOrder.setDateTime(qdt_end_of_today)

        myShop=config.get_shop_name("我的店铺")
        otherShop =config.get_shop_name("同行店铺")
        allShop = myShop+otherShop
        self.ui.cbOrderShop.setStyleSheet("QComboBox { combobox-popup: 0; } QComboBox QAbstractItemView { max-height: 400px; }")
        self.ui.cbOrderShop.addItems(allShop)

        #重置采购数据时间为当前时间
        self.ui.dateStart.setDate(QDate.currentDate())
        self.ui.dateEnd.setDate(QDate.currentDate())
        # 创建状态栏
        self.statusbar = self.statusBar()

        # 在状态栏中添加标签
        self.statusLabel = QLabel()
        self.statusbar.addWidget(self.statusLabel)




    #更新状态栏消息
    def update_status(self, message):
        self.statusLabel.setText(message)

    #导出店铺数据 btnExportOrder
    def on_order_export_click_listener(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Save File", "", "Excel Files (*.xlsx)")
        if file_path:
            wb = Workbook()
            ws = wb.active
            ws.title = "订单列表"

            # 写入表头
            for col in range(self.ui.tableOrder.columnCount()):
                header = self.ui.tableOrder.horizontalHeaderItem(col).text()
                ws.cell(row=1, column=col+1, value=header)

            # 写入数据
            for row in range(self.ui.tableOrder.rowCount()):
                for col in range(self.ui.tableOrder.columnCount()):
                    item = self.ui.tableOrder.item(row, col)
                    if item is not None:
                        ws.cell(row=row+2, column=col+1, value=item.text())

            wb.save(file_path)


    #店铺订单查询按钮
    def on_shop_order_click_listener(self):
        self.update_status("正在查询...")
        self.ui.tableOrder.setRowCount(0)
        #获取选择的店铺名称
        selectShop = self.ui.cbOrderShop.currentText()
        #获取开始和结束时间
        date_time_format = "yyyy-MM-dd HH:mm:ss"
        orderStart =self.ui.dataStartOrder.dateTime().toString(date_time_format)
        orderEnd =self.ui.dataEndOrder.dateTime().toString(date_time_format)
        self.requestData.request_order_data(selectShop,orderStart,orderEnd,self.order_data_load_call)

    #店铺下拉选择监听回调
    def selectionShopChanged(self,optoion):
        self.refresh_shop_data()
    
    def on_close(self):
        self.close()
    
    def on_mini(self):
        self.showMinimized()

    def on_tab_changed(self, index):
        self.update_status("")
        if index == 0:#店铺数据
            self.refresh_shop_data()
        elif index ==1:#采购数据
            self.refresh_purchase_data()
        elif index ==2:#店铺订单
            pass
        elif index ==3: #主图替换
            pass
        elif index ==4:#违规信息
            self.refresh_violation_data()
            pass

    #刷新店铺数据
    def refresh_shop_data(self):
        self.update_status("正在查询...")
        self.ui.tableWidget.setSortingEnabled(False)
        self.ui.tableWidget.setRowCount(0)
        selectedShop = self.ui.cbShop.currentText()
        self.requestData.request_shop_data(selectedShop,self.home_data_load_callback)


    #刷新采购数据
    def refresh_purchase_data(self):
        self.update_status("正在查询...")
        self.ui.purchaseTable.setSortingEnabled(False)
        self.ui.purchaseTable.setRowCount(0)
        isSelectPurchaseTime = self.ui.rbPurchase.isChecked()
        dateStart = self.ui.dateStart.date().toString(Qt.DateFormat.ISODate)
        dateEnd = self.ui.dateEnd.date().toString(Qt.DateFormat.ISODate)
        self.requestData.request_purchase_data(isSelectPurchaseTime,dateStart,dateEnd,self.data_load_purchase_callback)

    #刷新违规数据
    def refresh_violation_data(self):
        self.update_status("正在查询...")
        self.ui.treeViolation.clear()
        self.ui.treeTransaction.clear()
        self.requestData.request_violation_data(self.violation_data_load_call)
        self.requestData.request_transaction_data(self.transaction_data_load_call)
    
    #订单数据回调
    def order_data_load_call(self,dataList):
        #总刷单数量
        orderBrushingNumTotal = 0
        #刷单金额
        orderBrushingPriceTotal = 0.0
        #真实订单数量
        realOrderNumTotal =0
        #真实订单金额
        realOrderPriceTotal = 0.0
        for content in dataList:

            row_position = self.ui.tableOrder.rowCount()
            self.ui.tableOrder.insertRow(row_position)

            orderResult = content["orderResult"]

            warehouseGoodsList = content["warehouseGoodsList"]

            #购买的第一个商品
            warehouseGoods = warehouseGoodsList[0]
            b2COrderDetailResult = warehouseGoods["b2COrderDetailResultList"][0]


            #订单号
            orderViewId = orderResult["orderViewId"]
            orderViewIdItem = QtWidgets.QTableWidgetItem(str(orderViewId))
            self.ui.tableOrder.setItem(row_position, Contants.index_order_view, orderViewIdItem)


            #产品名称
            goodsName = b2COrderDetailResult["goodsName"]
            goodsNameItem = QtWidgets.QTableWidgetItem(str(goodsName))
            self.ui.tableOrder.setItem(row_position, Contants.index_order_goods_name, goodsNameItem)  

            # 下单时间
            orderTime = orderResult["orderTime"]
            # 将时间戳转换为 datetime 对象
            dtOrderTime = datetime.fromtimestamp(orderTime)
            # 格式化为字符串
            formattedOrderTime = dtOrderTime.strftime('%Y-%m-%d %H:%M:%S')
            orderTimeItem = QtWidgets.QTableWidgetItem(str(formattedOrderTime))
            self.ui.tableOrder.setItem(row_position, Contants.index_order_create_time, orderTimeItem) 


            #订单状态
            b2cOrderMainStatusDesc = orderResult["b2cOrderMainStatusDesc"]
            orderStatusItem = QtWidgets.QTableWidgetItem(str(b2cOrderMainStatusDesc))
            self.ui.tableOrder.setItem(row_position, Contants.index_order_status, orderStatusItem) 

            #购买数量
            warehouseGoodsCount = warehouseGoods["warehouseGoodsCount"]
            warehouseGoodsCountItem = QtWidgets.QTableWidgetItem(str(warehouseGoodsCount))
            self.ui.tableOrder.setItem(row_position, Contants.index_order_goods_count, warehouseGoodsCountItem) 

            #购买单价
            medicinePrice = b2COrderDetailResult["medicinePrice"]
            medicinePricetItem = QtWidgets.QTableWidgetItem(str(medicinePrice))
            self.ui.tableOrder.setItem(row_position, Contants.index_order_goods_price, medicinePricetItem) 

            #订单金额
            totalPrice = orderResult["total"]
            totalPriceItem = QtWidgets.QTableWidgetItem(str(totalPrice))
            self.ui.tableOrder.setItem(row_position, Contants.index_order_total_price, totalPriceItem) 

            #收货人
            recipientName = orderResult["recipientName"]
            recipientNameItem = QtWidgets.QTableWidgetItem(str(recipientName))
            self.ui.tableOrder.setItem(row_position, Contants.index_order_recipient_name, recipientNameItem) 

            #收获电话
            recipientPhone = orderResult["recipientPhone"]
            recipientPhoneItem = QtWidgets.QTableWidgetItem(str(recipientPhone))
            self.ui.tableOrder.setItem(row_position, Contants.index_order_recipient_phone, recipientPhoneItem) 

            #收获地址
            recipientAddress = orderResult["recipientAddress"]
            recipientAddressItem = QtWidgets.QTableWidgetItem(str(recipientAddress))
            self.ui.tableOrder.setItem(row_position, Contants.index_order_recipient_address, recipientAddressItem) 

            #下单渠道
            orderSource = orderResult["orderSource"]
            orderSourceItem = QtWidgets.QTableWidgetItem(str(orderSource))
            self.ui.tableOrder.setItem(row_position, Contants.index_order_recipient_source, orderSourceItem) 

            #是否刷单
            isBrushing = "否"
            if float(medicinePrice)<1.0 or warehouseGoodsCount>10:
                isBrushing = "疑似刷单"
                orderBrushingNumTotal+=1
                orderBrushingPriceTotal +=float(totalPrice)
            else:
                realOrderNumTotal+=1
                realOrderPriceTotal+=float(totalPrice)

            orderBrushingItem = QtWidgets.QTableWidgetItem(isBrushing)
            self.ui.tableOrder.setItem(row_position, Contants.index_order_brushing, orderBrushingItem) 

        #显示数量  
        self.ui.labelBottomOrfer.setText(f"总订单数：{len(dataList)} | 刷单数量：{orderBrushingNumTotal} | 刷单金额：{orderBrushingPriceTotal:.2f} | 真实单量{realOrderNumTotal} | 真实金额：{realOrderPriceTotal:.2f}")

         # 判断是否所有任务已经完成
        active_threads = self.requestData.threadpool.activeThreadCount()
        if active_threads == 0:
            # 启用排序
            self.update_status("查询完成")

 
    #店铺数据回调
    def home_data_load_callback(self,data,shopName):
        if data["code"] !=0:
            return
        shopData = data["data"]

        dealAmount = shopData["dealAmount"]
        dealAmountValue = dealAmount["amount"] #今日营业额
        dealRelativeCycle = dealAmount["relativeCycle"] #营业额环比
        dealRankingInterval = dealAmount["rankingInterval"] #成交额排名

        transactionVolume = shopData["transactionVolume"]
        transactionTurnover = transactionVolume["turnover"] #成交单量
        transactionRelativeCycle = transactionVolume["relativeCycle"] #单量环比


        singleAveragePrice = shopData["singleAveragePrice"]
        averagePrice = singleAveragePrice["averagePrice"] #单均价

        refundAmount = shopData["refundAmount"]
        refundAmountValue = refundAmount["amount"] #退款金额
        refundRelativeCycle = refundAmount["relativeCycle"] #退款金额环比

    
        row_position = self.ui.tableWidget.rowCount()
        self.ui.tableWidget.insertRow(row_position)
        
        #店铺名称
        shopName = QtWidgets.QTableWidgetItem(str(shopName))
        self.ui.tableWidget.setItem(row_position, Contants.index_name, shopName)

        #营业额
        dealAmountValueItem = NumericTableWidgetItem(str(dealAmountValue))
        self.ui.tableWidget.setItem(row_position, Contants.index_deal_amount_value, dealAmountValueItem)
        
        #营业额环比
        dealRelativeCycleItem = QtWidgets.QTableWidgetItem(str(dealRelativeCycle))
        if "-" in dealRelativeCycle: 
            dealRelativeCycleItem.setForeground(Qt.GlobalColor.green)
        else :
            dealRelativeCycleItem.setForeground(Qt.GlobalColor.red)
        self.ui.tableWidget.setItem(row_position, Contants.index_deal_relative_cycle, dealRelativeCycleItem)
        
        #营业额排名
        dealRankingIntervalItem = QtWidgets.QTableWidgetItem(str(dealRankingInterval))
        self.ui.tableWidget.setItem(row_position, Contants.index_deal_ranking_interval, dealRankingIntervalItem)
        
        #成交单量
        transactionTurnoverItem = QtWidgets.QTableWidgetItem(str(transactionTurnover))
        self.ui.tableWidget.setItem(row_position, Contants.index_transaction_turnover, transactionTurnoverItem)
        
        #单量环比
        transactionRelativeCycleItem = QtWidgets.QTableWidgetItem(str(transactionRelativeCycle))
        if "-" in transactionRelativeCycle: 
            transactionRelativeCycleItem.setForeground(Qt.GlobalColor.green)
        else :
            transactionRelativeCycleItem.setForeground(Qt.GlobalColor.red)
        self.ui.tableWidget.setItem(row_position, Contants.index_transaction_relative_cycle, transactionRelativeCycleItem)
              
        #单均价
        averagePriceItem = QtWidgets.QTableWidgetItem(str(averagePrice))
        self.ui.tableWidget.setItem(row_position, Contants.index_average_price, averagePriceItem)
        

        #退款金额
        refundAmountValueItem = QtWidgets.QTableWidgetItem(str(refundAmountValue))
        self.ui.tableWidget.setItem(row_position, Contants.index_refund_amount_value, refundAmountValueItem)
        

        #退款金额环比
        refundRelativeCycleItem = QtWidgets.QTableWidgetItem(str(refundRelativeCycle))
        if "-" in refundRelativeCycle: 
            refundRelativeCycleItem.setForeground(Qt.GlobalColor.green)
        else :
            refundRelativeCycleItem.setForeground(Qt.GlobalColor.red)

        self.ui.tableWidget.setItem(row_position, Contants.index_refund_relative_cycle, refundRelativeCycleItem)
        self.ui.tableWidget.update()
         # 判断是否所有任务已经完成
        active_threads = self.requestData.threadpool.activeThreadCount()
        if active_threads == 0:
            # 启用排序
            self.ui.tableWidget.setSortingEnabled(True)
            self.ui.tableWidget.sortItems(Contants.index_deal_amount_value, Qt.SortOrder.DescendingOrder)  # 按第二列（Amount 列）排序
            self.calculate_shop_total()
            self.update_status("查询完成")

    #经营违规回调
    def violation_data_load_call(self,dataJson):
        totalCount = dataJson["totalCount"]
        print("违规个数:"+str(totalCount))
        dataList = dataJson["data"]
        if len(dataList) == 0:
            return
        poiName = str(dataList[0]["poiName"]).replace("（快递电商）","")

        # 设置列标题
        self.ui.treeViolation.setHeaderLabels(["店铺名称", "处罚方式","日期"])
        #设置列宽
        self.ui.treeViolation.setColumnWidth(0,170)
        self.ui.treeViolation.setColumnWidth(1,140)
        self.ui.treeViolation.setColumnWidth(2,130)
        # 添加顶层项目
        parent1 = QTreeWidgetItem(self.ui.treeViolation, [poiName, "",""])

        # 添加子项目
        for item in dataList:
            toBTitle = item["toBTitle"]
            actionType =  item["actionTypes"][0]
            createTime = item["createTime"]
            # 转换为可视化日期和时间
            dt = datetime.fromtimestamp(createTime/1000.0)
            time = dt.strftime('%Y-%m-%d %H:%M:%S')
            child1 = QTreeWidgetItem(parent1, [toBTitle, actionType,time])
        
        # 判断是否所有任务已经完成
        active_threads = self.requestData.threadpool.activeThreadCount()
        if active_threads == 0:
            # 展开所有项
            self.ui.treeViolation.expandAll()
            self.update_status("查询完成")

    #交易风险违规
    def transaction_data_load_call(self,dataList):
        if len(dataList) ==0:
            return

        poiName = str(dataList[0]["wmPoiName"]).replace("（快递电商）","")
        # 设置列标题
        self.ui.treeTransaction.setHeaderLabels(["店铺名称", "处罚方式","日期"])
        #设置列宽
        self.ui.treeTransaction.setColumnWidth(0,170)
        self.ui.treeTransaction.setColumnWidth(1,140)
        self.ui.treeTransaction.setColumnWidth(2,130)
        # 添加顶层项目
        parent1 = QTreeWidgetItem(self.ui.treeTransaction, [poiName, "",""])

        for item in dataList:
            # 添加子项目
            violationTitle = item["violationTitle"]
            title =  item["punishInfo"][0]["title"]
            content =str(item["punishInfo"][0]["content"]).replace("时间：","")
            child1 = QTreeWidgetItem(parent1, [violationTitle, title,content])

        # 判断是否所有任务已经完成
        active_threads = self.requestData.threadpool2.activeThreadCount()
        if active_threads == 0:
            print("")
            # 展开所有项
            self.ui.treeTransaction.expandAll()
            self.update_status("查询完成")


    def calculate_shop_total(self):
        turnoverTotal = 0
        orderTotal = 0
        for row in range(self.ui.tableWidget.rowCount()):
            item = self.ui.tableWidget.item(row, Contants.index_deal_amount_value)
            if item is not None :
                turnoverTotal += float(item.text())

            item = self.ui.tableWidget.item(row, Contants.index_transaction_turnover)
            if item is not None :
                orderTotal += int(item.text())

        self.ui.amountCount.setText(f'总订单：{orderTotal}  营业额: {turnoverTotal:.2f}')

    #刷新采购数据
    def data_load_purchase_callback(self,dataJson,shopName):
        row_position = self.ui.purchaseTable.rowCount()
        self.ui.purchaseTable.insertRow(row_position)
        
        #店铺名称
        shopNameItem = QtWidgets.QTableWidgetItem(str(shopName))
        self.ui.purchaseTable.setItem(row_position, Contants.index_purchase_shop_name, shopNameItem)
        
        state = dataJson["success"]
        if not state:
            return
        data = dataJson["data"]

        pagination = data["pagination"]

        #订单总数
        total = pagination["total"]
        orderNumberItem = QtWidgets.QTableWidgetItem(str(total))
        self.ui.purchaseTable.setItem(row_position, Contants.index_purchase_order_number, orderNumberItem)
        
        #最新一条订单
        list = data["list"]
        if list == None or len(list)==0:
            return
        
        firstOrder = list[0]

        #采购时间
        addTime = firstOrder["addTime"]
        addTimeItem = QtWidgets.QTableWidgetItem(str(addTime))
        self.ui.purchaseTable.setItem(row_position, Contants.index_purchase_add_time, addTimeItem)

        #订单时间
        orderTime = firstOrder["orderTime"]
        orderTimeItem = QtWidgets.QTableWidgetItem(str(orderTime))
        self.ui.purchaseTable.setItem(row_position, Contants.index_purchase_order_time, orderTimeItem)

        paymentTotal = 0.0
        estimatedRevenueTotal = 0.0
        purchPaymentTotal = 0.0
        profitTotal = 0.0
        
        
        for item in list:
            paymentTotal+=item["payment"]
            estimatedRevenueTotal+=item["itemSellerPrice"]
            purchPaymentTotal+=item["purchPayment"]
            profitTotal+=item["profit"]

        #采购营业额
        paymentTotalItem = QtWidgets.QTableWidgetItem(f'{paymentTotal:.2f}')
        self.ui.purchaseTable.setItem(row_position, Contants.index_purchase_turnover, paymentTotalItem)
        #预计收入
        estimatedRevenueItem = QtWidgets.QTableWidgetItem(f'{estimatedRevenueTotal:.2f}')
        self.ui.purchaseTable.setItem(row_position, Contants.index_estimated_revenue, estimatedRevenueItem)
        #采购成本
        purchPaymentTotalItem = QtWidgets.QTableWidgetItem(f'{purchPaymentTotal:.2f}')
        self.ui.purchaseTable.setItem(row_position, Contants.index_purchase_cost, purchPaymentTotalItem)

        #采购利润
        profitTotalItem = NumericTableWidgetItem(f'{profitTotal:.2f}')
        profitTotalItem.setForeground(Qt.GlobalColor.red)
        self.ui.purchaseTable.setItem(row_position, Contants.index_purchase_profit, profitTotalItem)
        #利润率 
        profitMargin =  "{:.2f}%".format(profitTotal/estimatedRevenueTotal*100)
        profitMarginItem = QtWidgets.QTableWidgetItem(profitMargin)
        self.ui.purchaseTable.setItem(row_position, Contants.index_profit_margin, profitMarginItem)


         # 判断是否所有任务已经完成
        active_threads = self.requestData.threadpool.activeThreadCount()
        if active_threads == 0:
            # 启用排序
            self.ui.purchaseTable.setSortingEnabled(True)
            self.ui.purchaseTable.sortItems(Contants.index_purchase_profit, Qt.SortOrder.DescendingOrder) 
            self.calculate_profit_total()
            self.update_status("查询完成")

        

    def calculate_profit_total(self):
        orderTotal = 0
        profitTotal = 0
        for row in range(self.ui.purchaseTable.rowCount()):
            item = self.ui.purchaseTable.item(row, Contants.index_purchase_profit)
            if item is not None :
                profitTotal += float(item.text())
            item = self.ui.purchaseTable.item(row, Contants.index_purchase_order_number)
            if item is not None:
                orderTotal+=int(item.text())
        self.ui.labelPurchase.setText(f'总订单：{orderTotal}  总利润: {profitTotal:.2f}')


    def mousePressEvent(self, event):
        # 鼠标按下时记录坐标
        self.mouse_pos = event.globalPosition()
 
    def mouseMoveEvent(self, event):
        # 鼠标移动时计算移动距离
        if self.mouse_pos:
            diff = event.globalPosition() - self.mouse_pos
            self.move(self.pos() + diff.toPoint())
            self.mouse_pos = event.globalPosition()
 
    def mouseReleaseEvent(self, event):
        # 鼠标释放时清空坐标
        self.mouse_pos = None

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    # 设置全局字体大小
    app.setStyleSheet("""
        * {
            font-size: 14px;
        }
    """)
    mainWindow = MyApp()
    mainWindow.setWindowOpacity(0.9) # 设置窗口透明度
    mainWindow.setWindowFlag(Qt.WindowType.FramelessWindowHint)
    mainWindow.setStyleSheet(qdarkstyle.load_stylesheet())
    mainWindow.show()
    sys.exit(app.exec())